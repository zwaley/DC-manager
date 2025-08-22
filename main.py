import os
from fastapi import FastAPI, Request, Depends, Form, UploadFile, File, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session, aliased
from sqlalchemy import func, or_
import pandas as pd
from typing import List, Optional
from urllib.parse import quote
import io
import traceback # 导入 traceback 用于打印详细的错误堆栈
from datetime import datetime, timedelta, date
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pydantic import BaseModel

# 导入配置
from config import ADMIN_PASSWORD, PORT

# 修正了导入，使用正确的函数名和模型
from models import SessionLocal, Device, Connection, LifecycleRule, create_db_and_tables

def verify_admin_password(password: str) -> bool:
    """
    验证管理员密码
    Args:
        password: 用户输入的密码
    Returns:
        bool: 密码是否正确
    """
    return password == ADMIN_PASSWORD

# --- FastAPI 应用设置 ---

app = FastAPI(
    title="安吉电信动力设备管理系统",
    description="一个用于管理和可视化动力设备资产的Web应用。",
    version="1.1.0" # 版本升级
)

# 挂载静态文件目录
app.mount("/static", StaticFiles(directory="static"), name="static")
# 设置模板目录
templates = Jinja2Templates(directory="templates")

# --- 数据库会话管理 ---

def get_db():
    """
    数据库会话管理函数
    增加了详细的日志记录来跟踪数据库连接的创建和关闭过程
    """
    print("\n--- 创建数据库会话 ---")
    db = None
    try:
        db = SessionLocal()
        print(f"数据库会话创建成功: {id(db)}")
        yield db
    except Exception as e:
        print(f"数据库会话创建失败: {e}")
        if db:
            print("正在回滚数据库事务...")
            db.rollback()
        raise
    finally:
        if db:
            print(f"正在关闭数据库会话: {id(db)}")
            db.close()
            print("数据库会话已关闭")
        print("--- 数据库会话管理结束 ---\n")

# --- 应用启动事件 ---

@app.on_event("startup")
def on_startup():
    """
    应用启动事件处理函数
    增加了详细的日志记录来跟踪应用启动过程
    """
    print("\n" + "=" * 60)
    print("🚀 动力资源资产管理系统启动中...")
    print("=" * 60)
    
    try:
        # 检查并创建数据库目录
        db_dir = './database'
        if not os.path.exists(db_dir):
            print(f"📁 创建数据库目录: {db_dir}")
            os.makedirs(db_dir)
        else:
            print(f"📁 数据库目录已存在: {db_dir}")
        
        # 初始化数据库
        print("🗄️ 正在初始化数据库...")
        create_db_and_tables()
        
        print("✅ 应用启动完成！")
        print("🌐 服务器地址: http://localhost:8000")
        print("=" * 60 + "\n")
        
    except Exception as e:
        print(f"\n❌ 应用启动失败!")
        print(f"错误类型: {type(e).__name__}")
        print(f"错误信息: {e}")
        print("\n完整错误堆栈:")
        traceback.print_exc()
        print("=" * 60)
        raise  # 重新抛出异常，停止应用启动

# --- 路由和视图函数 ---

@app.get("/", response_class=HTMLResponse)
async def read_root(request: Request, db: Session = Depends(get_db)):
    """
    首页路由 - 显示所有设备列表
    增加了详细的日志记录来跟踪数据获取过程
    """
    print("\n=== 首页数据获取开始 ===")
    
    try:
        # 获取设备数据
        print("正在从数据库查询设备数据...")
        devices = db.query(Device).order_by(Device.id).all()
        device_count = len(devices)
        print(f"查询到 {device_count} 个设备")
        
        # 获取生命周期规则
        lifecycle_rules = db.query(LifecycleRule).filter(LifecycleRule.is_active == 'true').all()
        rules_dict = {rule.device_type: rule for rule in lifecycle_rules}
        print(f"加载了 {len(rules_dict)} 个生命周期规则")
        
        # 为每个设备计算生命周期状态
        for device in devices:
            lifecycle_status = "unknown"
            lifecycle_status_text = "未配置规则"
            
            if device.device_type and device.device_type in rules_dict:
                rule = rules_dict[device.device_type]
                if device.commission_date:
                    try:
                        # 解析投产日期
                        commission_date = None
                        date_str = str(device.commission_date).strip()
                        
                        # 处理特殊格式：YYYYMM (如 202312)
                        if re.match(r'^\d{6}$', date_str):
                            try:
                                year = int(date_str[:4])
                                month = int(date_str[4:6])
                                commission_date = datetime(year, month, 1)
                            except ValueError:
                                pass
                        
                        # 如果特殊格式解析失败，尝试标准格式
                        if not commission_date:
                            date_formats = [
                                "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d",
                                "%Y-%m", "%Y/%m", "%Y.%m",
                                "%Y年%m月%d日", "%Y年%m月"
                            ]
                            
                            for fmt in date_formats:
                                try:
                                    commission_date = datetime.strptime(date_str, fmt)
                                    break
                                except ValueError:
                                    continue
                        
                        if commission_date:
                            # 计算服役时间
                            today = datetime.now()
                            service_years = (today - commission_date).days / 365.25
                            
                            # 判断状态
                            if service_years >= rule.lifecycle_years:
                                lifecycle_status = "expired"
                                lifecycle_status_text = "已超期"
                            elif service_years >= (rule.lifecycle_years - rule.warning_months / 12):
                                lifecycle_status = "warning"
                                lifecycle_status_text = "临近超限"
                            else:
                                lifecycle_status = "normal"
                                lifecycle_status_text = "正常"
                        else:
                            lifecycle_status = "unknown"
                            lifecycle_status_text = "投产日期格式无法识别"
                    except Exception as e:
                        lifecycle_status = "unknown"
                        lifecycle_status_text = "投产日期格式无法识别"
                else:
                    lifecycle_status = "unknown"
                    lifecycle_status_text = "投产日期未填写"
            
            # 将状态信息添加到设备对象
            device.lifecycle_status = lifecycle_status
            device.lifecycle_status_text = lifecycle_status_text
        
        # 显示前几个设备的信息用于调试
        if device_count > 0:
            print("\n前3个设备信息:")
            for i, device in enumerate(devices[:3]):
                print(f"  设备{i+1}: ID={device.id}, 资产编号={device.asset_id}, 名称={device.name}, 生命周期状态={device.lifecycle_status}")
        else:
            print("警告: 数据库中没有设备数据！")
        
        # 获取连接数据用于统计
        connections = db.query(Connection).all()
        connection_count = len(connections)
        print(f"数据库中共有 {connection_count} 个连接")
        
        # 获取所有不重复的局站列表，用于筛选下拉框
        print("正在获取局站列表...")
        stations = db.query(Device.station).filter(Device.station.isnot(None)).filter(Device.station != '').distinct().all()
        station_list = [station[0] for station in stations if station[0]]  # 提取局站名称并过滤空值
        station_list.sort()  # 按字母顺序排序
        print(f"找到 {len(station_list)} 个不同的局站: {station_list}")
        
        # 获取所有不重复的设备类型列表，用于筛选下拉框
        print("正在获取设备类型列表...")
        device_types = db.query(Device.device_type).filter(Device.device_type.isnot(None)).filter(Device.device_type != '').distinct().all()
        device_type_list = [device_type[0] for device_type in device_types if device_type[0]]  # 提取设备类型并过滤空值
        device_type_list.sort()  # 按字母顺序排序
        print(f"找到 {len(device_type_list)} 个不同的设备类型: {device_type_list}")
        
        # 获取所有不重复的厂家列表，用于筛选下拉框
        print("正在获取厂家列表...")
        vendors = db.query(Device.vendor).filter(Device.vendor.isnot(None)).filter(Device.vendor != '').distinct().all()
        vendor_list = [vendor[0] for vendor in vendors if vendor[0]]  # 提取厂家名称并过滤空值
        vendor_list.sort()  # 按字母顺序排序
        print(f"找到 {len(vendor_list)} 个不同的厂家: {vendor_list}")
        
        # 检查是否有上传错误信息
        upload_error = request.query_params.get("error")
        if upload_error:
            print(f"检测到上传错误信息: {upload_error}")
        else:
            print("没有上传错误信息")
        
        # 检查是否有成功信息
        success_message = request.query_params.get("success")
        if success_message:
            print(f"检测到成功信息: {success_message}")
        else:
            print("没有成功信息")
        
        print("=== 首页数据获取完成 ===")
        
        return templates.TemplateResponse("index.html", {
            "request": request, 
            "devices": devices, 
            "stations": station_list,
            "device_types": device_type_list,
            "vendors": vendor_list,
            "upload_error": upload_error,
            "success_message": success_message
        })
        
    except Exception as e:
        print(f"\n!!! 首页数据获取失败 !!!")
        print(f"错误类型: {type(e).__name__}")
        print(f"错误信息: {e}")
        print("\n完整错误堆栈:")
        traceback.print_exc()
        print("=" * 50)
        
        # 返回错误页面或空设备列表
        return templates.TemplateResponse("index.html", {
            "request": request, 
            "devices": [], 
            "stations": [],
            "device_types": [],
            "vendors": [],
            "upload_error": f"获取设备数据时出错: {e}"
        })

@app.post("/upload")
async def upload_excel(file: UploadFile = File(...), password: str = Form(...), db: Session = Depends(get_db)):
    """
    处理 Excel 文件上传。
    如果失败，则重定向回主页并附带详细错误信息。
    增加了详细的日志记录来跟踪处理过程。
    """
    print("\n=== 开始处理上传的Excel文件 ===")
    print(f"上传文件名: {file.filename}")
    print(f"文件类型: {file.content_type}")
    
    # 验证管理员密码
    if not verify_admin_password(password):
        error_message = "密码错误，无权限执行此操作。"
        print(f"权限验证失败: {error_message}")
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)
    
    print("管理员密码验证通过")
    
    try:
        # 步骤 1: 增量更新模式 - 保留手工添加的设备，只更新Excel中的设备
        print("\n步骤 1: 采用增量更新模式，保留现有手工添加的设备...")
        
        # 记录当前数据量
        current_connections_count = db.query(Connection).count()
        current_devices_count = db.query(Device).count()
        print(f"当前数据库状态: {current_connections_count} 个连接, {current_devices_count} 个设备")
        print("步骤 1: 完成。将采用增量更新模式处理Excel数据。")

        contents = await file.read()
        print(f"文件大小: {len(contents)} 字节")
        buffer = io.BytesIO(contents)
        
        # 步骤 2: 读取Excel文件
        print("\n步骤 2: 使用 pandas 读取Excel文件...")
        # 通过 dtype 参数指定列以字符串形式读取，避免自动转换格式
        # 重要：假设"上级设备"列现在包含的是父设备的资产编号
        df = pd.read_excel(buffer, dtype={
            '资产编号': str,
            '设备投产时间': str,
            '上级设备': str 
        })
        df = df.where(pd.notna(df), None) # 将 NaN 替换为 None
        print(f"步骤 2: 完成。读取到 {len(df)} 行数据。")
        print(f"Excel 文件列名: {df.columns.tolist()}")
        
        # 验证必要的列是否存在
        required_columns = ['资产编号', '设备名称']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            error_msg = f"Excel文件缺少必要的列: {missing_columns}"
            print(f"错误: {error_msg}")
            return RedirectResponse(url=f"/?error={quote(error_msg)}", status_code=303)
        
        # 显示前几行数据样本用于调试
        print("\n前3行数据样本:")
        for i in range(min(3, len(df))):
            print(f"第{i+1}行: 资产编号={df.iloc[i].get('资产编号')}, 设备名称={df.iloc[i].get('设备名称')}")

        devices_map = {} # 这个映射将以 资产编号 为键
        devices_created_count = 0
        devices_updated_count = 0
        skipped_rows = []

        # 步骤 3: 增量更新设备（创建或更新）
        print("\n步骤 3: 开始第一遍处理 - 增量更新设备（创建新设备或更新现有设备）...")
        for index, row in df.iterrows():
            # 新增：获取并校验资产编号
            asset_id = row.get("资产编号")
            if isinstance(asset_id, str):
                asset_id = asset_id.strip()

            if not asset_id or asset_id == 'nan' or asset_id.lower() == 'none':
                skip_reason = f"资产编号为空或无效: '{asset_id}'"
                print(f"  - 第 {index+2} 行：跳过，{skip_reason}")
                skipped_rows.append((index+2, skip_reason))
                continue
            
            device_name = row.get("设备名称")
            if isinstance(device_name, str):
                device_name = device_name.strip()

            if not device_name or device_name == 'nan' or device_name.lower() == 'none':
                skip_reason = f"设备名称为空或无效: '{device_name}'"
                print(f"  - 第 {index+2} 行：跳过，{skip_reason}")
                skipped_rows.append((index+2, skip_reason))
                continue
            
            # 检查资产编号是否已在本次上传中重复
            if asset_id in devices_map:
                skip_reason = f"资产编号 '{asset_id}' 在Excel文件中重复"
                print(f"  - 第 {index+2} 行：跳过，{skip_reason}")
                skipped_rows.append((index+2, skip_reason))
                continue

            try:
                # 检查数据库中是否已存在该资产编号的设备
                existing_device = db.query(Device).filter(Device.asset_id == asset_id).first()
                
                # 获取局站信息
                station = row.get("局站")
                if isinstance(station, str):
                    station = station.strip()
                if not station or station == 'nan' or station.lower() == 'none':
                    skip_reason = f"局站信息为空或无效: '{station}'"
                    print(f"  - 第 {index+2} 行：跳过，{skip_reason}")
                    skipped_rows.append((index+2, skip_reason))
                    continue
                
                if existing_device:
                    # 更新现有设备
                    existing_device.name = device_name
                    existing_device.station = station
                    existing_device.model = row.get("设备型号")
                    existing_device.device_type = row.get("设备类型")
                    existing_device.location = row.get("机房内空间位置")
                    existing_device.power_rating = row.get("设备额定容量")
                    existing_device.vendor = row.get("设备生产厂家")
                    existing_device.commission_date = row.get("设备投产时间")
                    existing_device.remark = row.get("备注")
                    
                    devices_map[asset_id] = existing_device
                    devices_updated_count += 1
                    print(f"  - 第 {index+2} 行：准备更新现有设备 '{device_name}' (资产编号: {asset_id}, 局站: {station})")
                else:
                    # 创建新设备
                    device = Device(
                        asset_id=asset_id,
                        name=device_name,
                        station=station,
                        model=row.get("设备型号"),
                        device_type=row.get("设备类型"),
                        location=row.get("机房内空间位置"),
                        power_rating=row.get("设备额定容量"),
                        vendor=row.get("设备生产厂家"),
                        commission_date=row.get("设备投产时间"),
                        remark=row.get("备注")
                    )
                    db.add(device)
                    devices_map[asset_id] = device
                    devices_created_count += 1
                    print(f"  - 第 {index+2} 行：准备创建新设备 '{device_name}' (资产编号: {asset_id}, 局站: {station})")
                    
            except Exception as device_error:
                skip_reason = f"处理设备失败: {device_error}"
                print(f"  - 第 {index+2} 行：跳过，{skip_reason}")
                skipped_rows.append((index+2, skip_reason))
                continue
        
        print(f"\n准备提交设备更改到数据库（新建: {devices_created_count}, 更新: {devices_updated_count}）...")
        try:
            db.commit() # 提交事务以生成设备ID
            print("设备提交成功！")
        except Exception as commit_error:
            print(f"设备提交失败: {commit_error}")
            db.rollback()
            raise commit_error
            
        # 验证设备数量
        actual_device_count = db.query(Device).count()
        print(f"步骤 3: 完成。新建 {devices_created_count} 个设备，更新 {devices_updated_count} 个设备，数据库中总共有 {actual_device_count} 个设备。")
        
        if skipped_rows:
            print(f"\n跳过的行数统计: {len(skipped_rows)} 行")
            for row_num, reason in skipped_rows[:5]:  # 只显示前5个
                print(f"  第{row_num}行: {reason}")
            if len(skipped_rows) > 5:
                print(f"  ... 还有 {len(skipped_rows) - 5} 行被跳过")

        # 刷新映射，确保对象包含数据库生成的ID
        print("\n刷新设备对象以获取数据库生成的ID...")
        for asset_id_key in list(devices_map.keys()):
            try:
                db.refresh(devices_map[asset_id_key])
                print(f"  设备 {asset_id_key} ID: {devices_map[asset_id_key].id}")
            except Exception as refresh_error:
                print(f"  刷新设备 {asset_id_key} 失败: {refresh_error}")

        # 步骤 4: 清理涉及Excel设备的旧连接
        print("\n步骤 4: 清理涉及Excel中设备的旧连接...")
        excel_device_ids = [device.id for device in devices_map.values()]
        if excel_device_ids:
            # 删除涉及这些设备的所有连接（作为源设备或目标设备）
            old_connections_deleted = db.query(Connection).filter(
                (Connection.source_device_id.in_(excel_device_ids)) |
                (Connection.target_device_id.in_(excel_device_ids))
            ).delete(synchronize_session=False)
            db.commit()
            print(f"删除了 {old_connections_deleted} 个涉及Excel设备的旧连接")
        else:
            print("没有Excel设备，跳过连接清理")
            
        connections_created_count = 0
        connection_skipped_rows = []
        
        # 步骤 5: 创建新连接
        print("\n步骤 5: 开始第二遍处理 - 创建新连接...")
        for index, row in df.iterrows():
            # 使用资产编号来查找设备
            source_asset_id = row.get("上级设备")
            target_asset_id = row.get("资产编号")

            if isinstance(source_asset_id, str):
                source_asset_id = source_asset_id.strip()
            if isinstance(target_asset_id, str):
                target_asset_id = target_asset_id.strip()
            
            # 检查是否有上级设备信息
            if not source_asset_id or source_asset_id == 'nan' or source_asset_id.lower() == 'none':
                print(f"  - 第 {index+2} 行：跳过连接创建，无上级设备信息")
                continue
                
            # 确保源和目标设备都存在于映射中
            if target_asset_id and source_asset_id:
                if source_asset_id not in devices_map:
                    skip_reason = f"上级设备 '{source_asset_id}' 不存在"
                    print(f"  - 第 {index+2} 行：跳过连接，{skip_reason}")
                    connection_skipped_rows.append((index+2, skip_reason))
                    continue
                    
                if target_asset_id not in devices_map:
                    skip_reason = f"目标设备 '{target_asset_id}' 不存在"
                    print(f"  - 第 {index+2} 行：跳过连接，{skip_reason}")
                    connection_skipped_rows.append((index+2, skip_reason))
                    continue
                
                source_device = devices_map[source_asset_id]
                target_device = devices_map[target_asset_id]
                
                try:
                    connection = Connection(
                        source_device_id=source_device.id,
                        source_port=row.get("上级端口"),
                        target_device_id=target_device.id,
                        target_port=row.get("本端端口"),
                        cable_type=row.get("线缆类型")
                    )
                    db.add(connection)
                    connections_created_count += 1
                    print(f"  - 第 {index+2} 行：准备创建从 '{source_device.name}' 到 '{target_device.name}' 的连接")
                except Exception as conn_error:
                    skip_reason = f"创建连接对象失败: {conn_error}"
                    print(f"  - 第 {index+2} 行：跳过连接，{skip_reason}")
                    connection_skipped_rows.append((index+2, skip_reason))
                    continue
        
        print(f"\n准备提交 {connections_created_count} 个连接到数据库...")
        try:
            db.commit()
            print("连接提交成功！")
        except Exception as commit_error:
            print(f"连接提交失败: {commit_error}")
            db.rollback()
            raise commit_error
            
        # 验证连接是否真的被创建
        actual_connection_count = db.query(Connection).count()
        print(f"步骤 5: 完成。预期创建 {connections_created_count} 个连接，实际数据库中有 {actual_connection_count} 个连接。")
        
        if connection_skipped_rows:
            print(f"\n连接跳过的行数统计: {len(connection_skipped_rows)} 行")
            for row_num, reason in connection_skipped_rows[:5]:  # 只显示前5个
                print(f"  第{row_num}行: {reason}")
            if len(connection_skipped_rows) > 5:
                print(f"  ... 还有 {len(connection_skipped_rows) - 5} 行连接被跳过")
        
        # 步骤 6: 处理Sheet2连接数据
        sheet2_connections_count = 0
        sheet2_skipped_rows = []
        
        try:
            print("\n步骤 6: 开始处理Sheet2连接数据...")
            
            # 尝试读取Sheet2（连接表）
            try:
                # 重置buffer位置到开头，因为之前读取Sheet1时已经移动了位置
                buffer.seek(0)
                df_connections = pd.read_excel(buffer, sheet_name='连接')
                print(f"成功读取Sheet2，共 {len(df_connections)} 行连接数据")
            except Exception as sheet_error:
                print(f"无法读取Sheet2（连接表）: {sheet_error}")
                print("跳过Sheet2处理，继续完成导入")
                df_connections = None
            
            if df_connections is not None and len(df_connections) > 0:
                # 连接类型映射
                CONNECTION_TYPE_MAPPING = {
                    '电缆': 'cable',
                    '铜排': 'busbar', 
                    '母线': 'busway',
                    'cable': 'cable',
                    'busbar': 'busbar',
                    'busway': 'busway'
                }
                
                # 辅助函数：获取或创建设备
                def get_or_create_device(device_name: str, default_station: str = "未知站点"):
                    """获取设备，如果不存在则自动创建"""
                    if not device_name:
                        return None
                    
                    device = db.query(Device).filter(Device.name == device_name).first()
                    if not device:
                        # 自动创建设备
                        device = Device(
                            name=device_name,
                            asset_id=f"AUTO_{len(device_name)}_{hash(device_name) % 10000:04d}",  # 生成唯一资产编号
                            station=default_station,
                            device_type="待确认",
                            location="待确认",
                            remark="通过Excel导入时自动创建，请完善设备信息"
                        )
                        db.add(device)
                        db.flush()  # 获取ID但不提交
                        print(f"  * 自动创建设备: {device_name} (ID: {device.id})")
                    return device
                
                # 统计信息
                created_devices = []
                warnings = []
                
                for index, row in df_connections.iterrows():
                    try:
                        # 获取设备名称
                        source_device_name = str(row.get('A端设备名称', '')).strip()
                        target_device_name = str(row.get('B端设备名称', '')).strip()
                        
                        # 处理空设备名称的情况
                        if not source_device_name and not target_device_name:
                            skip_reason = "A端和B端设备名称都为空"
                            print(f"  - 第 {index+2} 行：跳过连接，{skip_reason}")
                            sheet2_skipped_rows.append((index+2, skip_reason))
                            continue
                        elif not source_device_name:
                            skip_reason = "A端设备名称为空"
                            print(f"  - 第 {index+2} 行：跳过连接，{skip_reason}")
                            sheet2_skipped_rows.append((index+2, skip_reason))
                            continue
                        elif not target_device_name:
                            skip_reason = "B端设备名称为空"
                            print(f"  - 第 {index+2} 行：跳过连接，{skip_reason}")
                            sheet2_skipped_rows.append((index+2, skip_reason))
                            continue
                        
                        # 获取或创建设备
                        source_device = get_or_create_device(source_device_name)
                        target_device = get_or_create_device(target_device_name)
                        
                        if not source_device or not target_device:
                            skip_reason = "设备创建失败"
                            print(f"  - 第 {index+2} 行：跳过连接，{skip_reason}")
                            sheet2_skipped_rows.append((index+2, skip_reason))
                            continue
                        
                        # 记录新创建的设备
                        if source_device.remark and "通过Excel导入时自动创建" in source_device.remark:
                            if source_device_name not in created_devices:
                                created_devices.append(source_device_name)
                        if target_device.remark and "通过Excel导入时自动创建" in target_device.remark:
                            if target_device_name not in created_devices:
                                created_devices.append(target_device_name)
                        
                        # 处理端口逻辑
                        def build_port_info(fuse_number, fuse_spec, breaker_number, breaker_spec):
                            """构建端口信息，优先使用熔丝，其次使用空开"""
                            fuse_num = str(fuse_number).strip() if pd.notna(fuse_number) else ''
                            fuse_sp = str(fuse_spec).strip() if pd.notna(fuse_spec) else ''
                            breaker_num = str(breaker_number).strip() if pd.notna(breaker_number) else ''
                            breaker_sp = str(breaker_spec).strip() if pd.notna(breaker_spec) else ''
                            
                            if fuse_num and fuse_num != 'nan':
                                return f"{fuse_num} ({fuse_sp})" if fuse_sp and fuse_sp != 'nan' else fuse_num
                            elif breaker_num and breaker_num != 'nan':
                                return f"{breaker_num} ({breaker_sp})" if breaker_sp and breaker_sp != 'nan' else breaker_num
                            else:
                                return None
                        
                        # 构建A端和B端端口信息
                        source_port = build_port_info(
                            row.get('A端熔丝编号'), row.get('A端熔丝规格'),
                            row.get('A端空开编号'), row.get('A端空开规格')
                        )
                        target_port = build_port_info(
                            row.get('B端熔丝编号'), row.get('B端熔丝规格'),
                            row.get('B端空开编号'), row.get('空开规格')
                        )
                        
                        # 处理连接类型
                        connection_type_raw = str(row.get('连接类型（电缆 / 铜排 / 母线）', 'cable')).strip()
                        connection_type = CONNECTION_TYPE_MAPPING.get(connection_type_raw, 'cable')
                        
                        # 检查是否已存在相同连接
                        existing_connection = db.query(Connection).filter(
                            Connection.source_device_id == source_device.id,
                            Connection.target_device_id == target_device.id,
                            Connection.source_port == source_port,
                            Connection.target_port == target_port
                        ).first()
                        
                        if existing_connection:
                            skip_reason = "连接已存在"
                            print(f"  - 第 {index+2} 行：跳过连接，{skip_reason}")
                            sheet2_skipped_rows.append((index+2, skip_reason))
                            continue
                        
                        # 创建连接对象
                        connection = Connection(
                            source_device_id=source_device.id,
                            target_device_id=target_device.id,
                            source_port=source_port,
                            target_port=target_port,
                            # A端信息
                            source_fuse_number=str(row.get('A端熔丝编号', '')).strip() if pd.notna(row.get('A端熔丝编号')) else None,
                            source_fuse_spec=str(row.get('A端熔丝规格', '')).strip() if pd.notna(row.get('A端熔丝规格')) else None,
                            source_breaker_number=str(row.get('A端空开编号', '')).strip() if pd.notna(row.get('A端空开编号')) else None,
                            source_breaker_spec=str(row.get('A端空开规格', '')).strip() if pd.notna(row.get('A端空开规格')) else None,
                            # B端信息
                            target_fuse_number=str(row.get('B端熔丝编号', '')).strip() if pd.notna(row.get('B端熔丝编号')) else None,
                            target_fuse_spec=str(row.get('B端熔丝规格', '')).strip() if pd.notna(row.get('B端熔丝规格')) else None,
                            target_breaker_number=str(row.get('B端空开编号', '')).strip() if pd.notna(row.get('B端空开编号')) else None,
                            target_breaker_spec=str(row.get('空开规格', '')).strip() if pd.notna(row.get('空开规格')) else None,
                            target_device_location=str(row.get('B端设备位置（非动力设备）', '')).strip() if pd.notna(row.get('B端设备位置（非动力设备）')) else None,
                            # 连接信息
                            hierarchy_relation=str(row.get('上下级', '')).strip() if pd.notna(row.get('上下级')) else None,
                            upstream_downstream=str(row.get('上下游', '')).strip() if pd.notna(row.get('上下游')) else None,
                            connection_type=connection_type,
                            cable_model=str(row.get('电缆型号', '')).strip() if pd.notna(row.get('电缆型号')) else None,
                            # 附加信息
                            source_device_photo=str(row.get('A端设备照片', '')).strip() if pd.notna(row.get('A端设备照片')) else None,
                            target_device_photo=str(row.get('B端设备照片', '')).strip() if pd.notna(row.get('B端设备照片')) else None,
                            remark=str(row.get('备注', '')).strip() if pd.notna(row.get('备注')) else None
                        )
                        
                        db.add(connection)
                        sheet2_connections_count += 1
                        print(f"  - 第 {index+2} 行：准备创建从 '{source_device_name}' 到 '{target_device_name}' 的连接")
                        print(f"    源端口: {source_port}, 目标端口: {target_port}, 连接类型: {connection_type}")
                        
                    except Exception as conn_error:
                        skip_reason = f"处理连接失败: {conn_error}"
                        print(f"  - 第 {index+2} 行：跳过连接，{skip_reason}")
                        sheet2_skipped_rows.append((index+2, skip_reason))
                        continue
                
                # 提交Sheet2连接
                if sheet2_connections_count > 0:
                    print(f"\n准备提交 {sheet2_connections_count} 个Sheet2连接到数据库...")
                    try:
                        db.commit()
                        print("Sheet2连接提交成功！")
                    except Exception as commit_error:
                        print(f"Sheet2连接提交失败: {commit_error}")
                        db.rollback()
                        raise commit_error
                
                # 生成详细的导入报告
                print(f"\n=== Sheet2连接导入报告 ===")
                print(f"总连接数: {len(df_connections)} 行")
                print(f"成功导入: {sheet2_connections_count} 个连接")
                print(f"跳过连接: {len(sheet2_skipped_rows)} 行")
                
                if created_devices:
                    print(f"\n自动创建的设备 ({len(created_devices)} 个):")
                    for device_name in created_devices:
                        print(f"  + {device_name}")
                    print("\n注意: 自动创建的设备信息不完整，请在设备管理页面完善相关信息。")
                
                if sheet2_skipped_rows:
                    print(f"\n跳过的连接详情:")
                    skip_reasons = {}
                    for row_num, reason in sheet2_skipped_rows:
                        if reason not in skip_reasons:
                            skip_reasons[reason] = []
                        skip_reasons[reason].append(row_num)
                    
                    for reason, rows in skip_reasons.items():
                        print(f"  {reason}: {len(rows)} 行 (第{', '.join(map(str, rows[:3]))}行{'...' if len(rows) > 3 else ''})")
                
                # 计算导入成功率
                success_rate = (sheet2_connections_count / len(df_connections)) * 100 if len(df_connections) > 0 else 0
                print(f"\n导入成功率: {success_rate:.1f}% ({sheet2_connections_count}/{len(df_connections)})")
            
            print(f"步骤 6: 完成。从Sheet2创建了 {sheet2_connections_count} 个连接")
            
        except Exception as sheet2_error:
            print(f"处理Sheet2时出错: {sheet2_error}")
            print("继续完成导入，忽略Sheet2错误")
        
        # 最终统计
        final_connection_count = db.query(Connection).count()
        total_connections_created = connections_created_count + sheet2_connections_count
        
        print("\n=== Excel文件增量更新处理成功 ===")
        print(f"处理结果: 新建 {devices_created_count} 个设备, 更新 {devices_updated_count} 个设备")
        print(f"连接创建: Sheet1创建 {connections_created_count} 个, Sheet2创建 {sheet2_connections_count} 个, 总计 {total_connections_created} 个")
        print(f"数据库最终状态: {actual_device_count} 个设备, {final_connection_count} 个连接")

    except Exception as e:
        print(f"\n!!! 发生异常，开始回滚事务 !!!")
        try:
            db.rollback()
            print("事务回滚成功")
        except Exception as rollback_error:
            print(f"事务回滚失败: {rollback_error}")
            
        error_message = f"处理Excel文件时出错: {e}"
        print(f"\n=== Excel文件处理失败 ===")
        print(f"错误类型: {type(e).__name__}")
        print(f"错误信息: {error_message}")
        print("\n完整错误堆栈:")
        traceback.print_exc()
        print("=" * 50)
        
        # 检查数据库状态
        try:
            final_device_count = db.query(Device).count()
            final_connection_count = db.query(Connection).count()
            print(f"\n错误后数据库状态: {final_device_count} 个设备, {final_connection_count} 个连接")
        except Exception as db_check_error:
            print(f"无法检查数据库状态: {db_check_error}")
            
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)

    print(f"\n上传处理完成，重定向到首页...")
    return RedirectResponse(url="/", status_code=303)

# 更新设备信息
@app.post("/devices/{device_id}")
async def update_device(
    device_id: int,
    asset_id: str = Form(...),
    name: str = Form(...),
    station: str = Form(...),
    model: str = Form(None),
    device_type: str = Form(None),
    location: str = Form(None),
    power_rating: str = Form(None),
    vendor: str = Form(None),
    commission_date: str = Form(None),
    remark: str = Form(None),
    db: Session = Depends(get_db)
):
    """更新设备信息（编辑功能不需要密码验证，因为在进入编辑页面时已验证）"""
    try:
        # 获取要更新的设备
        device = db.query(Device).filter(Device.id == device_id).first()
        if not device:
            error_message = "设备不存在。"
            return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)
        
        # 检查资产编号唯一性（排除当前设备）
        existing_device = db.query(Device).filter(
            Device.asset_id == asset_id,
            Device.id != device_id
        ).first()
        if existing_device:
            error_message = f"资产编号 {asset_id} 已存在，请使用其他编号。"
            return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)
        
        # 更新设备信息
        device.asset_id = asset_id
        device.name = name
        device.station = station
        device.model = model if model else None
        device.device_type = device_type if device_type else None
        device.location = location if location else None
        device.power_rating = power_rating if power_rating else None
        device.vendor = vendor if vendor else None
        device.commission_date = commission_date if commission_date else None
        device.remark = remark if remark else None
        
        db.commit()
        
        success_message = f"设备 {name} 更新成功。"
        return RedirectResponse(url=f"/?success={quote(success_message)}", status_code=303)
        
    except Exception as e:
        db.rollback()
        error_message = f"更新设备失败：{str(e)}"
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)

# 编辑设备页面
@app.get("/edit/{device_id}")
async def edit_device_page(device_id: int, password: str, request: Request, db: Session = Depends(get_db)):
    """显示编辑设备页面"""
    # 验证管理员密码
    if not verify_admin_password(password):
        error_message = "密码错误，无权限执行此操作。"
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)
    
    # 获取设备信息
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        error_message = "设备不存在。"
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)
    
    return templates.TemplateResponse("edit_device.html", {
        "request": request,
        "device": device
    })

# 删除设备
@app.delete("/devices/{device_id}")
async def delete_device(device_id: int, request: Request, db: Session = Depends(get_db)):
    """删除设备"""
    try:
        # 获取请求体中的密码
        body = await request.json()
        password = body.get("password")
        
        # 验证管理员密码
        if not verify_admin_password(password):
            raise HTTPException(status_code=403, detail="密码错误，无权限执行此操作。")
        
        # 获取要删除的设备
        device = db.query(Device).filter(Device.id == device_id).first()
        if not device:
            raise HTTPException(status_code=404, detail="设备不存在。")
        
        device_name = device.name
        
        # 删除相关的连接记录
        db.query(Connection).filter(
            (Connection.source_device_id == device_id) | 
            (Connection.target_device_id == device_id)
        ).delete()
        
        # 删除设备
        db.delete(device)
        db.commit()
        
        return {"message": f"设备 {device_name} 删除成功。"}
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"删除设备失败：{str(e)}")

@app.post("/devices")
async def create_device(
    asset_id: str = Form(...),
    name: str = Form(...),
    station: str = Form(...),
    model: str = Form(None),
    device_type: str = Form(None),
    location: str = Form(None),
    power_rating: str = Form(None),
    vendor: str = Form(None),
    commission_date: str = Form(None),
    remark: str = Form(None),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    # 验证管理员密码
    if not verify_admin_password(password):
        error_message = "密码错误，无权限执行此操作。"
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)
    
    # 增加资产编号唯一性校验
    existing_device = db.query(Device).filter(Device.asset_id == asset_id).first()
    if existing_device:
        # 如果存在，则重定向回主页并显示错误信息
        error_message = f"创建失败：资产编号 '{asset_id}' 已存在。"
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)

    new_device = Device(
        asset_id=asset_id,
        name=name,
        station=station,
        model=model,
        device_type=device_type,
        location=location,
        power_rating=power_rating,
        vendor=vendor,
        commission_date=commission_date,
        remark=remark
    )
    db.add(new_device)
    db.commit()
    return RedirectResponse(url="/", status_code=303)

@app.get("/graph_data/{device_id}")
async def get_graph_data(device_id: int, db: Session = Depends(get_db)):
    nodes = []
    edges = []
    processed_ids = set()

    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")

    queue = [device]
    visited_ids = {device.id}

    while queue:
        current_device = queue.pop(0)

        if current_device.id not in processed_ids:
            # 在悬浮提示中也加入资产编号
            nodes.append({
                "id": current_device.id,
                "label": current_device.name,
                "title": f"""<b>资产编号:</b> {current_device.asset_id}<br>
                             <b>名称:</b> {current_device.name}<br>
                             <b>型号:</b> {current_device.model or 'N/A'}<br>
                             <b>位置:</b> {current_device.location or 'N/A'}<br>
                             <b>功率:</b> {current_device.power_rating or 'N/A'}""",
                "level": 0 
            })
            processed_ids.add(current_device.id)

        # 向上游查找
        for conn in current_device.target_connections:
            source_device = conn.source_device
            if source_device and source_device.id not in visited_ids:
                edges.append({"from": source_device.id, "to": current_device.id, "arrows": "to", "label": conn.cable_type or ""})
                visited_ids.add(source_device.id)
                queue.append(source_device)

        # 向下游查找
        for conn in current_device.source_connections:
            target_device = conn.target_device
            if target_device and target_device.id not in visited_ids:
                edges.append({"from": current_device.id, "to": target_device.id, "arrows": "to", "label": conn.cable_type or ""})
                visited_ids.add(target_device.id)
                queue.append(target_device)
                
    return JSONResponse(content={"nodes": nodes, "edges": edges})


# 新增API路径：/api/power-chain/{device_id} - 与/graph_data/{device_id}功能相同，保持向后兼容
@app.get("/api/power-chain/{device_id}")
async def get_power_chain_data(device_id: int, db: Session = Depends(get_db)):
    """获取设备电力链路拓扑图数据 - 新的API路径
    
    Args:
        device_id: 设备ID
        db: 数据库会话
        
    Returns:
        JSONResponse: 包含nodes和edges的拓扑图数据
    """
    nodes = []
    edges = []
    processed_ids = set()

    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")

    queue = [device]
    visited_ids = {device.id}

    while queue:
        current_device = queue.pop(0)

        if current_device.id not in processed_ids:
            # 在悬浮提示中也加入资产编号
            nodes.append({
                "id": current_device.id,
                "label": current_device.name,
                "title": f"""<b>资产编号:</b> {current_device.asset_id}<br>
                             <b>名称:</b> {current_device.name}<br>
                             <b>型号:</b> {current_device.model or 'N/A'}<br>
                             <b>位置:</b> {current_device.location or 'N/A'}<br>
                             <b>功率:</b> {current_device.power_rating or 'N/A'}""",
                "level": 0 
            })
            processed_ids.add(current_device.id)

        # 向上游查找
        for conn in current_device.target_connections:
            source_device = conn.source_device
            if source_device and source_device.id not in visited_ids:
                edges.append({"from": source_device.id, "to": current_device.id, "arrows": "to", "label": conn.cable_type or ""})
                visited_ids.add(source_device.id)
                queue.append(source_device)

        # 向下游查找
        for conn in current_device.source_connections:
            target_device = conn.target_device
            if target_device and target_device.id not in visited_ids:
                edges.append({"from": current_device.id, "to": target_device.id, "arrows": "to", "label": conn.cable_type or ""})
                visited_ids.add(target_device.id)
                queue.append(target_device)
                
    return JSONResponse(content={"nodes": nodes, "edges": edges})


@app.get("/graph/{device_id}", response_class=HTMLResponse)
async def get_power_chain_graph(request: Request, device_id: int):
    return templates.TemplateResponse("graph.html", {"request": request, "device_id": device_id})


# --- 设备生命周期规则管理 API ---

@app.get("/api/lifecycle-rules")
async def get_lifecycle_rules(db: Session = Depends(get_db)):
    """
    获取所有生命周期规则
    """
    try:
        rules = db.query(LifecycleRule).all()
        return JSONResponse(content={
            "success": True,
            "data": [{
                "id": rule.id,
                "device_type": rule.device_type,
                "lifecycle_years": rule.lifecycle_years,
                "warning_months": rule.warning_months,
                "description": rule.description,
                "is_active": rule.is_active,
                "created_at": rule.created_at,
                "updated_at": rule.updated_at
            } for rule in rules]
        })
    except Exception as e:
        print(f"获取生命周期规则失败: {e}")
        return JSONResponse(content={"success": False, "message": str(e)}, status_code=500)


@app.post("/api/lifecycle-rules")
async def create_lifecycle_rule(
    device_type: str = Form(...),
    lifecycle_years: int = Form(...),
    warning_months: int = Form(6),
    description: str = Form(""),
    password: str = Form(...),  # 添加密码参数
    db: Session = Depends(get_db)
):
    """
    创建生命周期规则
    """
    try:
        # 验证管理员密码
        if not verify_admin_password(password):
            return JSONResponse(content={"success": False, "message": "密码错误"}, status_code=401)
        
        from datetime import datetime
        
        # 检查设备类型是否已存在规则
        existing_rule = db.query(LifecycleRule).filter(LifecycleRule.device_type == device_type).first()
        if existing_rule:
            return JSONResponse(content={
                "success": False, 
                "message": f"设备类型 '{device_type}' 的生命周期规则已存在"
            }, status_code=400)
        
        # 创建新规则
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_rule = LifecycleRule(
            device_type=device_type,
            lifecycle_years=lifecycle_years,
            warning_months=warning_months,
            description=description,
            is_active="true",
            created_at=current_time,
            updated_at=current_time
        )
        
        db.add(new_rule)
        db.commit()
        db.refresh(new_rule)
        
        return JSONResponse(content={
            "success": True,
            "message": "生命周期规则创建成功",
            "data": {
                "id": new_rule.id,
                "device_type": new_rule.device_type,
                "lifecycle_years": new_rule.lifecycle_years,
                "warning_months": new_rule.warning_months
            }
        })
        
    except Exception as e:
        db.rollback()
        print(f"创建生命周期规则失败: {e}")
        return JSONResponse(content={"success": False, "message": str(e)}, status_code=500)


@app.put("/api/lifecycle-rules/{rule_id}")
async def update_lifecycle_rule(
    rule_id: int,
    device_type: str = Form(...),
    lifecycle_years: int = Form(...),
    warning_months: int = Form(6),
    description: str = Form(""),
    is_active: str = Form("true"),
    password: str = Form(...),  # 添加密码参数
    db: Session = Depends(get_db)
):
    """
    更新生命周期规则
    """
    try:
        # 验证管理员密码
        if not verify_admin_password(password):
            return JSONResponse(content={"success": False, "message": "密码错误"}, status_code=401)
        
        from datetime import datetime
        
        rule = db.query(LifecycleRule).filter(LifecycleRule.id == rule_id).first()
        if not rule:
            return JSONResponse(content={"success": False, "message": "规则不存在"}, status_code=404)
        
        # 检查设备类型是否与其他规则冲突
        existing_rule = db.query(LifecycleRule).filter(
            LifecycleRule.device_type == device_type,
            LifecycleRule.id != rule_id
        ).first()
        if existing_rule:
            return JSONResponse(content={
                "success": False, 
                "message": f"设备类型 '{device_type}' 的生命周期规则已存在"
            }, status_code=400)
        
        # 更新规则
        rule.device_type = device_type
        rule.lifecycle_years = lifecycle_years
        rule.warning_months = warning_months
        rule.description = description
        rule.is_active = is_active
        rule.updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        db.commit()
        
        return JSONResponse(content={
            "success": True,
            "message": "生命周期规则更新成功"
        })
        
    except Exception as e:
        db.rollback()
        print(f"更新生命周期规则失败: {e}")
        return JSONResponse(content={"success": False, "message": str(e)}, status_code=500)


@app.delete("/api/lifecycle-rules/{rule_id}")
async def delete_lifecycle_rule(rule_id: int, password: str = Form(...), db: Session = Depends(get_db)):
    """
    删除生命周期规则
    """
    try:
        # 验证管理员密码
        if not verify_admin_password(password):
            return JSONResponse(content={"success": False, "message": "密码错误"}, status_code=401)
        
        rule = db.query(LifecycleRule).filter(LifecycleRule.id == rule_id).first()
        if not rule:
            return JSONResponse(content={"success": False, "message": "规则不存在"}, status_code=404)
        
        db.delete(rule)
        db.commit()
        
        return JSONResponse(content={
            "success": True,
            "message": "生命周期规则删除成功"
        })
        
    except Exception as e:
        db.rollback()
        print(f"删除生命周期规则失败: {e}")
        return JSONResponse(content={"success": False, "message": str(e)}, status_code=500)


@app.get("/api/devices")
async def get_devices_api(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(50, ge=1, le=200, description="每页数量"),
    db: Session = Depends(get_db)
):
    """
    获取设备列表API接口
    """
    try:
        # 构建查询
        query = db.query(Device)
        
        # 计算总数
        total = query.count()
        
        # 应用分页
        offset = (page - 1) * page_size
        devices = query.offset(offset).limit(page_size).all()
        
        # 构建响应数据
        result = []
        for device in devices:
            result.append({
                "id": device.id,
                "asset_id": device.asset_id,
                "name": device.name,
                "station": device.station,
                "model": device.model,
                "device_type": device.device_type,
                "location": device.location,
                "power_rating": device.power_rating,
                "vendor": device.vendor,
                "commission_date": device.commission_date.isoformat() if device.commission_date and hasattr(device.commission_date, 'isoformat') else device.commission_date,
                "remark": device.remark
            })
        
        return JSONResponse(content={
            "success": True,
            "data": result,
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total": total,
                "total_pages": (total + page_size - 1) // page_size
            }
        })
        
    except Exception as e:
        print(f"获取设备列表失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取设备列表失败: {str(e)}")


@app.get("/api/devices/lifecycle-status")
async def get_devices_lifecycle_status(
    status_filter: Optional[str] = None,  # normal, warning, expired, all
    db: Session = Depends(get_db)
):
    """
    获取设备生命周期状态
    status_filter: normal(正常), warning(临近超限), expired(已超期), all(全部)
    """
    try:
        from datetime import datetime, timedelta
        import re
        
        # 获取所有设备和规则
        devices = db.query(Device).all()
        rules = {rule.device_type: rule for rule in db.query(LifecycleRule).filter(LifecycleRule.is_active == "true").all()}
        
        result_devices = []
        current_date = datetime.now()
        
        for device in devices:
            # 查找对应的生命周期规则
            rule = rules.get(device.device_type)
            if not rule:
                # 没有规则的设备标记为未知状态
                device_info = {
                    "id": device.id,
                    "asset_id": device.asset_id,
                    "name": device.name,
                    "station": device.station,
                    "model": device.model,
                    "vendor": device.vendor,
                    "commission_date": device.commission_date,
                    "lifecycle_status": "unknown",
                    "lifecycle_status_text": "未配置规则",
                    "days_in_service": None,
                    "remaining_days": None,
                    "rule_years": None
                }
                if not status_filter or status_filter == "all":
                    result_devices.append(device_info)
                continue
            
            # 解析投产日期
            if not device.commission_date:
                device_info = {
                    "id": device.id,
                    "asset_id": device.asset_id,
                    "name": device.name,
                    "station": device.station,
                    "model": device.model,
                    "vendor": device.vendor,
                    "commission_date": device.commission_date,
                    "lifecycle_status": "unknown",
                    "lifecycle_status_text": "投产日期未填写",
                    "days_in_service": None,
                    "remaining_days": None,
                    "rule_years": rule.lifecycle_years
                }
                if not status_filter or status_filter == "all":
                    result_devices.append(device_info)
                continue
            
            # 尝试解析多种日期格式
            commission_date = None
            date_str = device.commission_date.strip()
            
            # 处理特殊格式：YYYYMM (如 202312)
            if re.match(r'^\d{6}$', date_str):
                try:
                    year = int(date_str[:4])
                    month = int(date_str[4:6])
                    commission_date = datetime(year, month, 1)
                except ValueError:
                    pass
            
            # 如果特殊格式解析失败，尝试标准格式
            if not commission_date:
                date_formats = [
                    "%Y-%m-%d",
                    "%Y/%m/%d", 
                    "%Y.%m.%d",
                    "%Y-%m",
                    "%Y/%m",
                    "%Y.%m",
                    "%Y"
                ]
                
                for fmt in date_formats:
                    try:
                        if fmt == "%Y":
                            # 只有年份的情况，默认为该年的1月1日
                            commission_date = datetime.strptime(device.commission_date, fmt).replace(month=1, day=1)
                        elif fmt in ["%Y-%m", "%Y/%m", "%Y.%m"]:
                            # 只有年月的情况，默认为该月的1日
                            commission_date = datetime.strptime(device.commission_date, fmt).replace(day=1)
                        else:
                            commission_date = datetime.strptime(device.commission_date, fmt)
                        break
                    except ValueError:
                        continue
            
            if not commission_date:
                device_info = {
                    "id": device.id,
                    "asset_id": device.asset_id,
                    "name": device.name,
                    "station": device.station,
                    "model": device.model,
                    "vendor": device.vendor,
                    "commission_date": device.commission_date,
                    "lifecycle_status": "unknown",
                    "lifecycle_status_text": "投产日期格式无法识别",
                    "days_in_service": None,
                    "remaining_days": None,
                    "rule_years": rule.lifecycle_years
                }
                if not status_filter or status_filter == "all":
                    result_devices.append(device_info)
                continue
            
            # 计算服役时间和剩余时间
            days_in_service = (current_date - commission_date).days
            lifecycle_days = rule.lifecycle_years * 365
            remaining_days = lifecycle_days - days_in_service
            warning_days = rule.warning_months * 30
            
            # 确定生命周期状态
            if remaining_days < 0:
                lifecycle_status = "expired"
                lifecycle_status_text = f"已超期 {abs(remaining_days)} 天"
            elif remaining_days <= warning_days:
                lifecycle_status = "warning"
                lifecycle_status_text = f"临近超限，剩余 {remaining_days} 天"
            else:
                lifecycle_status = "normal"
                lifecycle_status_text = f"正常，剩余 {remaining_days} 天"
            
            device_info = {
                "id": device.id,
                "asset_id": device.asset_id,
                "name": device.name,
                "station": device.station,
                "model": device.model,
                "vendor": device.vendor,
                "commission_date": device.commission_date,
                "lifecycle_status": lifecycle_status,
                "lifecycle_status_text": lifecycle_status_text,
                "days_in_service": days_in_service,
                "remaining_days": remaining_days,
                "rule_years": rule.lifecycle_years
            }
            
            # 根据筛选条件添加设备
            if not status_filter or status_filter == "all" or status_filter == lifecycle_status:
                result_devices.append(device_info)
        
        # 统计信息
        total_count = len(result_devices)
        normal_count = len([d for d in result_devices if d["lifecycle_status"] == "normal"])
        warning_count = len([d for d in result_devices if d["lifecycle_status"] == "warning"])
        expired_count = len([d for d in result_devices if d["lifecycle_status"] == "expired"])
        unknown_count = len([d for d in result_devices if d["lifecycle_status"] == "unknown"])
        
        return JSONResponse(content={
            "success": True,
            "data": result_devices,
            "statistics": {
                "total": total_count,
                "normal": normal_count,
                "warning": warning_count,
                "expired": expired_count,
                "unknown": unknown_count
            }
        })
        
    except Exception as e:
        print(f"获取设备生命周期状态失败: {e}")
        traceback.print_exc()
        return JSONResponse(content={"success": False, "message": str(e)}, status_code=500)


@app.get("/test-route")
async def test_route():
    """
    测试路由
    """
    print("=== 测试路由被调用 ===")
    return {"message": "测试路由正常工作", "timestamp": "updated"}

@app.get("/debug-routes")
async def debug_routes():
    """
    调试路由 - 显示所有已注册的路由
    """
    routes = []
    for route in app.routes:
        if hasattr(route, 'path') and hasattr(route, 'methods'):
            routes.append({
                "path": route.path,
                "methods": list(route.methods) if route.methods else [],
                "name": getattr(route, 'name', 'unknown')
            })
    return {"registered_routes": routes, "total_count": len(routes)}

@app.get("/debug-lifecycle")
async def debug_lifecycle():
    """
    调试生命周期路由
    """
    print("=== 调试生命周期路由被调用 ===")
    return {"message": "调试路由正常工作", "status": "ok"}

@app.post("/api/verify-password")
async def verify_password(request: Request):
    """
    验证管理员密码
    """
    try:
        data = await request.json()
        password = data.get("password", "")
        
        if verify_admin_password(password):
            return {"success": True, "message": "密码验证成功"}
        else:
            return {"success": False, "message": "密码错误"}
    except Exception as e:
        print(f"Error verifying password: {e}")
        return {"success": False, "message": "验证失败"}

@app.get("/lifecycle-management", response_class=HTMLResponse)
async def lifecycle_management_page(request: Request):
    """
    生命周期管理页面
    """
    print("=== 访问生命周期管理页面 ===")
    print(f"请求URL: {request.url}")
    print(f"请求方法: {request.method}")
    try:
        print("正在渲染模板...")
        response = templates.TemplateResponse("lifecycle_management.html", {"request": request})
        print("模板渲染成功")
        return response
    except Exception as e:
        print(f"生命周期管理页面错误: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/connections", response_class=HTMLResponse)
async def connections_page(request: Request):
    """
    连接管理页面
    """
    print("=== 访问连接管理页面 ===")
    print(f"请求URL: {request.url}")
    print(f"请求方法: {request.method}")
    try:
        print("正在渲染模板...")
        response = templates.TemplateResponse("connections.html", {"request": request})
        print("模板渲染成功")
        return response
    except Exception as e:
        print(f"连接管理页面错误: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/export")
async def export_devices(
    password: str = Form(...),
    export_range: str = Form("all"),
    station_filter: str = Form(""),
    name_filter: str = Form(""),
    device_type_filter: str = Form(""),
    vendor_filter: str = Form(""),
    lifecycle_filter: str = Form(""),
    db: Session = Depends(get_db)
):
    """
    导出设备数据为Excel文件
    支持全量导出和筛选导出，需要管理员密码验证
    """
    try:
        # 验证管理员密码
        if not verify_admin_password(password):
            raise HTTPException(status_code=401, detail="密码错误，无权限导出数据")
        
        # 根据导出范围查询设备数据
        query = db.query(Device)
        
        # 如果是筛选导出，应用筛选条件
        if export_range == "filtered":
            if station_filter:
                query = query.filter(Device.station.ilike(f"%{station_filter}%"))
            if name_filter:
                query = query.filter(Device.name.ilike(f"%{name_filter}%"))
            if device_type_filter:
                query = query.filter(Device.device_type.ilike(f"%{device_type_filter}%"))
            if vendor_filter:
                query = query.filter(Device.vendor.ilike(f"%{vendor_filter}%"))
            if lifecycle_filter:
                # 这里需要根据生命周期状态筛选，暂时跳过复杂的生命周期逻辑
                pass
        
        devices = query.all()
        
        if not devices:
            raise HTTPException(status_code=404, detail="没有找到设备数据")
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "设备列表"
        
        # 定义表头
        headers = [
            "ID", "资产编号", "设备名称", "局站", "设备类型", "设备型号", 
            "所在位置", "额定容量", "设备生产厂家", "投产日期", "备注"
        ]
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 写入设备数据
        for row, device in enumerate(devices, 2):
            data = [
                device.id,
                device.asset_id,
                device.name,
                device.station,
                device.device_type,
                device.model,
                device.location,
                device.power_rating,
                device.vendor,
                device.commission_date.strftime("%Y-%m-%d") if device.commission_date else "",
                device.remark
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # 设置斑马纹效果
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 限制最大宽度
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 冻结首行
        ws.freeze_panes = "A2"
        
        # 添加筛选器
        ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}1"
        
        # 生成文件名（包含时间戳）
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if export_range == "filtered":
            filename = f"设备列表_筛选导出_{timestamp}.xlsx"
        else:
            filename = f"设备列表_全量导出_{timestamp}.xlsx"
        
        # 将Excel文件保存到内存
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # 设置响应头
        headers = {
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        
        # 返回文件流
        return StreamingResponse(
            io.BytesIO(excel_buffer.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"导出设备数据错误: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


# --- 连接管理 Pydantic 模型 ---

class ConnectionCreate(BaseModel):
    """创建连接的请求模型"""
    source_device_id: int
    target_device_id: int
    connection_type: Optional[str] = None
    cable_model: Optional[str] = None
    source_fuse_number: Optional[str] = None
    source_fuse_spec: Optional[str] = None
    source_breaker_number: Optional[str] = None
    source_breaker_spec: Optional[str] = None
    target_fuse_number: Optional[str] = None
    target_fuse_spec: Optional[str] = None
    target_breaker_number: Optional[str] = None
    target_breaker_spec: Optional[str] = None
    hierarchy_relation: Optional[str] = None
    upstream_downstream: Optional[str] = None
    parallel_count: Optional[int] = 1
    rated_current: Optional[float] = None
    cable_length: Optional[float] = None
    source_device_photo: Optional[str] = None
    target_device_photo: Optional[str] = None
    remark: Optional[str] = None
    installation_date: Optional[date] = None

class ConnectionUpdate(BaseModel):
    """更新连接的请求模型"""
    source_device_id: Optional[int] = None
    target_device_id: Optional[int] = None
    connection_type: Optional[str] = None
    cable_model: Optional[str] = None
    source_fuse_number: Optional[str] = None
    source_fuse_spec: Optional[str] = None
    source_breaker_number: Optional[str] = None
    source_breaker_spec: Optional[str] = None
    target_fuse_number: Optional[str] = None
    target_fuse_spec: Optional[str] = None
    target_breaker_number: Optional[str] = None
    target_breaker_spec: Optional[str] = None
    hierarchy_relation: Optional[str] = None
    upstream_downstream: Optional[str] = None
    parallel_count: Optional[int] = None
    rated_current: Optional[float] = None
    cable_length: Optional[float] = None
    source_device_photo: Optional[str] = None
    target_device_photo: Optional[str] = None
    remark: Optional[str] = None
    installation_date: Optional[date] = None

class ConnectionResponse(BaseModel):
    """连接响应模型"""
    id: int
    source_device_id: int
    target_device_id: int
    source_device_name: str
    target_device_name: str
    source_port: Optional[str]  # 源端口名称（带前缀）
    target_port: Optional[str]  # 目标端口名称（带前缀）
    connection_type: Optional[str]
    cable_model: Optional[str]
    source_fuse_number: Optional[str]
    source_fuse_spec: Optional[str]
    source_breaker_number: Optional[str]
    source_breaker_spec: Optional[str]
    target_fuse_number: Optional[str]
    target_fuse_spec: Optional[str]
    target_breaker_number: Optional[str]
    target_breaker_spec: Optional[str]
    hierarchy_relation: Optional[str]
    upstream_downstream: Optional[str]
    parallel_count: Optional[int]
    rated_current: Optional[float]
    cable_length: Optional[float]
    source_device_photo: Optional[str]
    target_device_photo: Optional[str]
    remark: Optional[str]
    installation_date: Optional[date]
    created_at: datetime
    updated_at: Optional[datetime]
    
    class Config:
        # 启用ORM模式，允许从SQLAlchemy模型创建
        from_attributes = True
        # 自定义JSON编码器处理日期时间对象
        json_encoders = {
            datetime: lambda v: v.isoformat() if v else None,
            date: lambda v: v.isoformat() if v else None
        }


# --- 连接管理 RESTful API 接口 ---

@app.get("/api/connections/statistics")
async def get_connections_statistics(db: Session = Depends(get_db)):
    """
    获取连接统计信息
    """
    try:
        # 总连接数
        total_connections = db.query(Connection).count()
        
        # 按连接类型统计
        connection_type_stats = db.query(
            Connection.connection_type,
            func.count(Connection.id).label('count')
        ).group_by(Connection.connection_type).all()
        
        # 将混合的中英文连接类型统计合并为标准格式
        cable_count = 0
        busbar_count = 0
        bus_count = 0
        
        for item in connection_type_stats:
            conn_type = item[0] or ""
            count = item[1]
            
            # 电缆类型（cable 或 电缆）
            if conn_type.lower() in ['cable', '电缆']:
                cable_count += count
            # 铜排类型（busbar 或 铜排）
            elif conn_type.lower() in ['busbar', '铜排']:
                busbar_count += count
            # 母线类型（bus、busway 或 母线）
            elif conn_type.lower() in ['bus', 'busway', '母线']:
                bus_count += count
        
        # 按设备类型统计（源设备）
        device_type_stats = db.query(
            Device.device_type,
            func.count(Connection.id).label('count')
        ).join(Connection, Device.id == Connection.source_device_id)\
         .group_by(Device.device_type).all()
        
        # 最近30天新增连接数
        thirty_days_ago = datetime.now() - timedelta(days=30)
        recent_connections = db.query(Connection)\
            .filter(Connection.created_at >= thirty_days_ago).count()
        
        return JSONResponse(content={
            "success": True,
            "data": {
                "total": total_connections,
                "cable": cable_count,
                "busbar": busbar_count,
                "bus": bus_count,
                "recent_connections": recent_connections,
                "connection_types": [
                    {"type": item[0] or "未分类", "count": item[1]} 
                    for item in connection_type_stats
                ],
                "device_types": [
                    {"type": item[0] or "未分类", "count": item[1]} 
                    for item in device_type_stats
                ]
            }
        })
        
    except Exception as e:
        print(f"获取连接统计失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取连接统计失败: {str(e)}")


@app.get("/api/connections")
async def get_connections(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(100, ge=1, le=5000, description="每页数量"),
    source_device_id: Optional[int] = Query(None, description="源设备ID"),
    target_device_id: Optional[int] = Query(None, description="目标设备ID"),
    connection_type: Optional[str] = Query(None, description="连接类型"),
    device_name: Optional[str] = Query(None, description="设备名称（模糊查询，匹配源设备或目标设备）"),
    db: Session = Depends(get_db)
):
    """
    获取连接列表
    支持分页和筛选功能
    """
    try:
        # 构建查询
        # 创建Device表的别名用于目标设备
        target_device = aliased(Device)
        query = db.query(Connection, Device.name.label('source_device_name'), target_device.name.label('target_device_name'))\
                  .join(Device, Connection.source_device_id == Device.id)\
                  .join(target_device, Connection.target_device_id == target_device.id)
        
        # 应用筛选条件
        if source_device_id:
            query = query.filter(Connection.source_device_id == source_device_id)
        if target_device_id:
            query = query.filter(Connection.target_device_id == target_device_id)
        if connection_type:
            query = query.filter(Connection.connection_type.ilike(f"%{connection_type}%"))
        # 按设备名称模糊查询（匹配源设备或目标设备）
        if device_name:
            query = query.filter(
                or_(
                    Device.name.ilike(f"%{device_name}%"),  # 匹配源设备名称
                    target_device.name.ilike(f"%{device_name}%")  # 匹配目标设备名称
                )
            )
        
        # 计算总数
        total = query.count()
        
        # 应用分页
        offset = (page - 1) * page_size
        results = query.offset(offset).limit(page_size).all()
        
        # 辅助函数：根据熔丝/空开编号为端口名称添加前缀
        def build_port_name_with_prefix(fuse_number, breaker_number, original_port=None):
            """根据熔丝编号或空开编号为端口名称添加前缀"""
            fuse_num = str(fuse_number).strip() if fuse_number and str(fuse_number).strip() not in ['', 'nan', 'None'] else ''
            breaker_num = str(breaker_number).strip() if breaker_number and str(breaker_number).strip() not in ['', 'nan', 'None'] else ''
            
            # 优先使用熔丝编号
            if fuse_num:
                return f"熔丝_{fuse_num}"
            elif breaker_num:
                return f"空开_{breaker_num}"
            else:
                # 如果都没有，返回原始端口名称或空字符串
                return original_port if original_port else ''
        
        # 构建响应数据 - 手动序列化日期字段以避免JSON序列化错误
        result = []
        for conn, source_name, target_name in results:
            # 手动处理日期字段的序列化
            installation_date_str = conn.installation_date.isoformat() if conn.installation_date else None
            created_at_str = conn.created_at.isoformat() if conn.created_at else None
            updated_at_str = conn.updated_at.isoformat() if conn.updated_at else None
            
            # 构建带前缀的端口名称
            source_port_with_prefix = build_port_name_with_prefix(
                conn.source_fuse_number, 
                conn.source_breaker_number, 
                conn.source_port
            )
            target_port_with_prefix = build_port_name_with_prefix(
                conn.target_fuse_number, 
                conn.target_breaker_number, 
                conn.target_port
            )
            
            result.append({
                "id": conn.id,
                "source_device_id": conn.source_device_id,
                "target_device_id": conn.target_device_id,
                "source_device_name": source_name,
                "target_device_name": target_name,
                "connection_type": conn.connection_type,
                "cable_model": conn.cable_model,
                "source_port": source_port_with_prefix,  # 使用带前缀的端口名称
                "target_port": target_port_with_prefix,  # 使用带前缀的端口名称
                "source_fuse_number": conn.source_fuse_number,
                "source_fuse_spec": conn.source_fuse_spec,
                "source_breaker_number": conn.source_breaker_number,
                "source_breaker_spec": conn.source_breaker_spec,
                "target_fuse_number": conn.target_fuse_number,
                "target_fuse_spec": conn.target_fuse_spec,
                "target_breaker_number": conn.target_breaker_number,
                "target_breaker_spec": conn.target_breaker_spec,
                "hierarchy_relation": conn.hierarchy_relation,
                "upstream_downstream": conn.upstream_downstream,
                "parallel_count": conn.parallel_count,
                "rated_current": conn.rated_current,
                "cable_length": conn.cable_length,
                "source_device_photo": conn.source_device_photo,
                "target_device_photo": conn.target_device_photo,
                "remark": conn.remark,
                "installation_date": installation_date_str,
                "created_at": created_at_str,
                "updated_at": updated_at_str
            })

        return {
            "success": True,
            "data": result,
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total": total,
                "total_pages": (total + page_size - 1) // page_size
            }
        }
        
    except Exception as e:
        print(f"获取连接列表失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取连接列表失败: {str(e)}")


@app.post("/api/connections")
async def create_connection(
    source_device_id: int = Form(...),
    target_device_id: int = Form(...),
    connection_type: Optional[str] = Form(None),
    cable_model: Optional[str] = Form(None),
    source_port: Optional[str] = Form(None),
    target_port: Optional[str] = Form(None),
    source_fuse_number: Optional[str] = Form(None),
    source_fuse_spec: Optional[str] = Form(None),
    source_breaker_number: Optional[str] = Form(None),
    source_breaker_spec: Optional[str] = Form(None),
    target_fuse_number: Optional[str] = Form(None),
    target_fuse_spec: Optional[str] = Form(None),
    target_breaker_number: Optional[str] = Form(None),
    target_breaker_spec: Optional[str] = Form(None),
    hierarchy_relation: Optional[str] = Form(None),
    upstream_downstream: Optional[str] = Form(None),
    parallel_count: Optional[int] = Form(1),
    rated_current: Optional[float] = Form(None),
    cable_length: Optional[float] = Form(None),
    source_device_photo: Optional[str] = Form(None),
    target_device_photo: Optional[str] = Form(None),
    remark: Optional[str] = Form(None),
    installation_date: Optional[str] = Form(None),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    """
    创建新连接
    需要管理员密码验证
    """
    try:
        # 验证管理员密码
        if not verify_admin_password(password):
            raise HTTPException(status_code=401, detail="密码错误")
        
        # 处理日期字段 - 支持yyyymm格式
        parsed_installation_date = None
        if installation_date:
            try:
                # 支持yyyymm格式，如202412
                if len(installation_date) == 6 and installation_date.isdigit():
                    year = int(installation_date[:4])
                    month = int(installation_date[4:6])
                    parsed_installation_date = datetime(year, month, 1).date()
                else:
                    raise ValueError("日期格式不正确")
            except ValueError:
                raise HTTPException(status_code=400, detail="安装日期格式错误，请使用YYYYMM格式（如：202412）")
        
        # 验证源设备和目标设备是否存在
        source_device = db.query(Device).filter(Device.id == source_device_id).first()
        if not source_device:
            raise HTTPException(status_code=404, detail=f"源设备ID {source_device_id} 不存在")
        
        target_device = db.query(Device).filter(Device.id == target_device_id).first()
        if not target_device:
            raise HTTPException(status_code=404, detail=f"目标设备ID {target_device_id} 不存在")
        
        # 检查是否已存在相同的连接
        existing_connection = db.query(Connection).filter(
            Connection.source_device_id == source_device_id,
            Connection.target_device_id == target_device_id
        ).first()
        
        if existing_connection:
            raise HTTPException(status_code=400, detail="该连接已存在")
        
        # 创建新连接
        new_connection = Connection(
            source_device_id=source_device_id,
            target_device_id=target_device_id,
            source_port=source_port,
            target_port=target_port,
            connection_type=connection_type,
            cable_model=cable_model,
            source_fuse_number=source_fuse_number,
            source_fuse_spec=source_fuse_spec,
            source_breaker_number=source_breaker_number,
            source_breaker_spec=source_breaker_spec,
            target_fuse_number=target_fuse_number,
            target_fuse_spec=target_fuse_spec,
            target_breaker_number=target_breaker_number,
            target_breaker_spec=target_breaker_spec,
            hierarchy_relation=hierarchy_relation,
            upstream_downstream=upstream_downstream,
            parallel_count=parallel_count,
            rated_current=rated_current,
            cable_length=cable_length,
            source_device_photo=source_device_photo,
            target_device_photo=target_device_photo,
            remark=remark,
            installation_date=parsed_installation_date,
            created_at=datetime.now()
        )
        
        db.add(new_connection)
        db.commit()
        db.refresh(new_connection)
        
        # 构建响应
        response = ConnectionResponse(
            id=new_connection.id,
            source_device_id=new_connection.source_device_id,
            target_device_id=new_connection.target_device_id,
            source_device_name=source_device.name,
            target_device_name=target_device.name,
            connection_type=new_connection.connection_type,
            cable_model=new_connection.cable_model,
            source_fuse_number=new_connection.source_fuse_number,
            source_fuse_spec=new_connection.source_fuse_spec,
            source_breaker_number=new_connection.source_breaker_number,
            source_breaker_spec=new_connection.source_breaker_spec,
            target_fuse_number=new_connection.target_fuse_number,
            target_fuse_spec=new_connection.target_fuse_spec,
            target_breaker_number=new_connection.target_breaker_number,
            target_breaker_spec=new_connection.target_breaker_spec,
            hierarchy_relation=new_connection.hierarchy_relation,
            upstream_downstream=new_connection.upstream_downstream,
            parallel_count=new_connection.parallel_count,
            rated_current=new_connection.rated_current,
            cable_length=new_connection.cable_length,
            source_device_photo=new_connection.source_device_photo,
            target_device_photo=new_connection.target_device_photo,
            remark=new_connection.remark,
            installation_date=new_connection.installation_date,
            created_at=new_connection.created_at,
            updated_at=new_connection.updated_at
        )
        
        # 手动处理日期字段序列化
        response_data = {
            "id": new_connection.id,
            "source_device_id": new_connection.source_device_id,
            "target_device_id": new_connection.target_device_id,
            "source_device_name": source_device.name,
            "target_device_name": target_device.name,
            "connection_type": new_connection.connection_type,
            "cable_model": new_connection.cable_model,
            "source_fuse_number": new_connection.source_fuse_number,
            "source_fuse_spec": new_connection.source_fuse_spec,
            "source_breaker_number": new_connection.source_breaker_number,
            "source_breaker_spec": new_connection.source_breaker_spec,
            "target_fuse_number": new_connection.target_fuse_number,
            "target_fuse_spec": new_connection.target_fuse_spec,
            "target_breaker_number": new_connection.target_breaker_number,
            "target_breaker_spec": new_connection.target_breaker_spec,
            "hierarchy_relation": new_connection.hierarchy_relation,
            "upstream_downstream": new_connection.upstream_downstream,
            "parallel_count": new_connection.parallel_count,
            "rated_current": new_connection.rated_current,
            "cable_length": new_connection.cable_length,
            "source_device_photo": new_connection.source_device_photo,
            "target_device_photo": new_connection.target_device_photo,
            "remark": new_connection.remark,
            "installation_date": new_connection.installation_date.isoformat() if new_connection.installation_date else None,
            "created_at": new_connection.created_at.isoformat() if new_connection.created_at else None,
            "updated_at": new_connection.updated_at.isoformat() if new_connection.updated_at else None
        }
        
        return JSONResponse(content={
            "success": True,
            "message": "连接创建成功",
            "data": response_data
        })
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"创建连接失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"创建连接失败: {str(e)}")


@app.put("/api/connections/{connection_id}", response_model=ConnectionResponse)
async def update_connection(
    connection_id: int,
    connection: ConnectionUpdate,
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    """
    更新连接信息
    需要管理员密码验证
    """
    try:
        # 验证管理员密码
        if not verify_admin_password(password):
            raise HTTPException(status_code=401, detail="密码错误")
        
        # 查找要更新的连接
        existing_connection = db.query(Connection).filter(Connection.id == connection_id).first()
        if not existing_connection:
            raise HTTPException(status_code=404, detail="连接不存在")
        
        # 如果要更新设备ID，验证设备是否存在
        if connection.source_device_id is not None:
            source_device = db.query(Device).filter(Device.id == connection.source_device_id).first()
            if not source_device:
                raise HTTPException(status_code=404, detail=f"源设备ID {connection.source_device_id} 不存在")
            existing_connection.source_device_id = connection.source_device_id
        
        if connection.target_device_id is not None:
            target_device = db.query(Device).filter(Device.id == connection.target_device_id).first()
            if not target_device:
                raise HTTPException(status_code=404, detail=f"目标设备ID {connection.target_device_id} 不存在")
            existing_connection.target_device_id = connection.target_device_id
        
        # 更新其他字段
        update_data = connection.dict(exclude_unset=True)
        for field, value in update_data.items():
            if field not in ['source_device_id', 'target_device_id']:  # 这两个字段已经处理过了
                setattr(existing_connection, field, value)
        
        existing_connection.updated_at = datetime.now()
        
        db.commit()
        db.refresh(existing_connection)
        
        # 构建响应
        response = ConnectionResponse(
            id=existing_connection.id,
            source_device_id=existing_connection.source_device_id,
            target_device_id=existing_connection.target_device_id,
            source_device_name=existing_connection.source_device.name,
            target_device_name=existing_connection.target_device.name,
            connection_type=existing_connection.connection_type,
            cable_model=existing_connection.cable_model,
            source_fuse_number=existing_connection.source_fuse_number,
            source_fuse_spec=existing_connection.source_fuse_spec,
            source_breaker_number=existing_connection.source_breaker_number,
            source_breaker_spec=existing_connection.source_breaker_spec,
            target_fuse_number=existing_connection.target_fuse_number,
            target_fuse_spec=existing_connection.target_fuse_spec,
            target_breaker_number=existing_connection.target_breaker_number,
            target_breaker_spec=existing_connection.target_breaker_spec,
            hierarchy_relation=existing_connection.hierarchy_relation,
            upstream_downstream=existing_connection.upstream_downstream,
            parallel_count=existing_connection.parallel_count,
            rated_current=existing_connection.rated_current,
            cable_length=existing_connection.cable_length,
            source_device_photo=existing_connection.source_device_photo,
            target_device_photo=existing_connection.target_device_photo,
            remark=existing_connection.remark,
            installation_date=existing_connection.installation_date,
            created_at=existing_connection.created_at,
            updated_at=existing_connection.updated_at
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "连接更新成功",
            "data": response.dict()
        })
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"更新连接失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"更新连接失败: {str(e)}")


@app.delete("/api/connections/{connection_id}")
async def delete_connection(
    connection_id: int,
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    """
    删除连接
    需要管理员密码验证
    """
    try:
        # 验证管理员密码
        if not verify_admin_password(password):
            raise HTTPException(status_code=401, detail="密码错误")
        
        # 查找要删除的连接
        connection = db.query(Connection).filter(Connection.id == connection_id).first()
        if not connection:
            raise HTTPException(status_code=404, detail="连接不存在")
        
        # 删除连接
        db.delete(connection)
        db.commit()
        
        return JSONResponse(content={
            "success": True,
            "message": "连接删除成功"
        })
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"删除连接失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"删除连接失败: {str(e)}")


@app.get("/api/connections/{connection_id}", response_model=ConnectionResponse)
async def get_connection(
    connection_id: int,
    db: Session = Depends(get_db)
):
    """
    获取单个连接详情
    """
    try:
        connection = db.query(Connection).filter(Connection.id == connection_id).first()
        if not connection:
            raise HTTPException(status_code=404, detail="连接不存在")
        
        response = ConnectionResponse(
            id=connection.id,
            source_device_id=connection.source_device_id,
            target_device_id=connection.target_device_id,
            source_device_name=connection.source_device.name,
            target_device_name=connection.target_device.name,
            source_port=build_port_name_with_prefix(
                connection.source_fuse_number, 
                connection.source_breaker_number
            ),
            target_port=build_port_name_with_prefix(
                connection.target_fuse_number, 
                connection.target_breaker_number
            ),
            connection_type=connection.connection_type,
            cable_model=connection.cable_model,
            source_fuse_number=connection.source_fuse_number,
            source_fuse_spec=connection.source_fuse_spec,
            source_breaker_number=connection.source_breaker_number,
            source_breaker_spec=connection.source_breaker_spec,
            target_fuse_number=connection.target_fuse_number,
            target_fuse_spec=connection.target_fuse_spec,
            target_breaker_number=connection.target_breaker_number,
            target_breaker_spec=connection.target_breaker_spec,
            hierarchy_relation=connection.hierarchy_relation,
            upstream_downstream=connection.upstream_downstream,
            parallel_count=connection.parallel_count,
            rated_current=connection.rated_current,
            cable_length=connection.cable_length,
            source_device_photo=connection.source_device_photo,
            target_device_photo=connection.target_device_photo,
            remark=connection.remark,
            installation_date=connection.installation_date,
            created_at=connection.created_at,
            updated_at=connection.updated_at
        )
        
        return JSONResponse(content={
            "success": True,
            "data": response.dict()
        })
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"获取连接详情失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取连接详情失败: {str(e)}")

# --- 应用启动 ---
if __name__ == "__main__":
    import uvicorn
    print(f"\n🌐 服务器启动地址: http://localhost:{PORT}")
    print(f"📊 管理界面: http://localhost:{PORT}")
    print(f"🔗 连接管理: http://localhost:{PORT}/connections")
    print(f"⚙️  生命周期管理: http://localhost:{PORT}/lifecycle-management")
    print("=" * 60)
    uvicorn.run(app, host="0.0.0.0", port=PORT, reload=False)